import io
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import barcode
from barcode.writer import ImageWriter
import qrcode
from openpyxl import load_workbook
from datetime import datetime
import os
import threading
import time
import win32print
import win32api
import tempfile
import pywintypes
import win32con
import win32ui
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 手动定义可能缺失的常量
if not hasattr(win32con, 'DM_SIZE'):
    win32con.DM_SIZE = 40  # DEVMODE结构的大小，适用于Windows XP及以上系统
if not hasattr(win32con, 'DMPAPER_USER'):
    win32con.DMPAPER_USER = 0  # 用户自定义纸张


    
class CompleteBarcodeGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("条码生成与打印工具")
        self.root.geometry("1100x750")
        self.root.minsize(1000, 600)

        # 打印状态相关变量
        self.printing = False
        self.blink_id = None  # 用于存储闪烁定时器ID

        # 设置样式
        self.style = ttk.Style()
        self.style.configure("Card.TFrame", background="#f0f0f0", borderwidth=1, relief="solid", padding=10)
        self.style.configure("Title.TLabel", font=('Arial', 10, 'bold'), background="#f0f0f0")
        self.style.configure("Normal.TLabel", font=('Arial', 9), background="#f0f0f0")

        # 公司名称变量
        self.company_name = tk.StringVar()
        self.company_name.set("珠海XXXXXX技术有限公司")

        # 打印机名称变量
        self.printer_name = tk.StringVar()
        self.available_printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
        if self.available_printers:
            self.printer_name.set(self.available_printers[0][2])

        # 创建界面元素
        self.create_widgets()

    def create_widgets(self):
        # 顶部控制面板
        control_frame = tk.Frame(self.root, bg="#f0f0f0", padx=10, pady=10)
        control_frame.pack(fill="x")

        # 文件和条码操作按钮
        tk.Button(control_frame, text="选择 Excel 文件", command=self.load_excel,
                  bg="#4CAF50", fg="white", relief="flat").pack(side="left", padx=5)
        tk.Button(control_frame, text="生成条码", command=self.generate_barcodes,
                  bg="#2196F3", fg="white", relief="flat").pack(side="left", padx=5)
        tk.Button(control_frame, text="清除", command=self.clear_display,
                  bg="#f44336", fg="white", relief="flat").pack(side="left", padx=5)
        tk.Button(control_frame, text="更新条码", command=self.update_barcodes,
                  bg="#FF9800", fg="white", relief="flat").pack(side="left", padx=5)
        # 打印操作按钮
        printer_menu = ttk.OptionMenu(control_frame, self.printer_name, *[printer[2] for printer in self.available_printers])
        printer_menu.pack(side="left", padx=5)
        tk.Button(control_frame, text="打印当前", command=self.print_current,
                  bg="#00BCD4", fg="white", relief="flat").pack(side="left", padx=5)
        tk.Button(control_frame, text="打印所有", command=self.print_all,
                  bg="#E91E63", fg="white", relief="flat").pack(side="left", padx=5)

        # 状态指示灯
        status_indicators = tk.Frame(control_frame, bg="#f0f0f0")
        status_indicators.pack(side="right", padx=10)

        tk.Label(status_indicators, text="打印状态:", bg="#f0f0f0").pack(side="left")
        self.print_indicator = tk.Label(status_indicators, text="", width=5, bg="red", relief="solid")
        self.print_indicator.pack(side="left")

        # 状态栏
        self.status_label = tk.Label(self.root, text="准备就绪", bd=1, relief="sunken",
                                     anchor="w", bg="#f0f0f0", fg="#333")
        self.status_label.pack(fill="x", padx=10, pady=(0, 5))

        # 主显示区域
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # 创建 Canvas 和 Scrollbar
        self.canvas = tk.Canvas(self.main_frame, bg="#f5f5f5")
        self.scrollbar = ttk.Scrollbar(self.main_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # 初始化变量
        self.file_path = ""
        self.workbook = None
        self.entry_widgets = []  # 存储输入框引用
        self.barcode_images = []  # 存储条码图像引用
        self.qrcode_images = []  # 存储二维码图像引用
    def print_text_info(self, hDC, x_pos, y_pos, fields):
        for label, value in fields:
            text = f"{label}: {value}"
            hDC.TextOut(x_pos, y_pos, text)
            y_pos += 20 * 1440 / 72
        return y_pos

    def print_image_info(self, hDC, x_pos, y_pos, img_labels):
        for img_label in img_labels:
            if img_label[0]:
                img = img_label[1]
                img = img.convert("RGB")
                photo = ImageTk.PhotoImage(img)
                img_width, img_height = img.size
                hDC.StretchBlt(x_pos, y_pos, img_width * 1440 / 72, img_height * 1440 / 72,
                            photo.width(), photo.height(), 0, 0, win32con.SRCCOPY)
                y_pos += img_height * 1440 / 72 + 20 * 1440 / 72
        return y_pos
    def print_current(self):
        """打印当前选中的条码信息（包含条码和二维码）"""
        if not self.entry_widgets:
            messagebox.showwarning("警告", "没有可打印的条码")
            return

        # 开始打印流程
        self.printing = True
        self._start_blink()
        self.status_label.config(text="正在打印...")

        # 在后台线程中执行打印
        threading.Thread(target=self._print_current_with_barcodes, daemon=True).start()
    def print_image(self):
        """打印当前项目的所有信息"""
        try:
            # 获取默认打印机
            printer_name = win32print.GetDefaultPrinter()
            hDC = win32ui.CreateDC()
            hDC.CreatePrinterDC(printer_name)
            # 设置映射模式为TWIPS
            hDC.SetMapMode(win32con.MM_TWIPS)
            
            # 开始打印作业
            hDC.StartDoc("Barcode and Information Print")
            hDC.StartPage()
            
            # 遍历当前显示的项目信息
            for row in self.entry_widgets:
                x_pos = 100  # 初始x坐标，单位为twips
                y_pos = 100  # 初始y坐标，单位为twips
                fields = [
                    ("PO", row[0].get()),
                    ("日期", row[1].get()),
                    ("项目号", row[2].get()),
                    ("料号", row[3].get()),
                    ("数量", row[4].get()),
                    ("名称", row[5].get())
                ]
                y_pos = self.print_text_info(hDC, x_pos, y_pos, fields)
                y_pos = self.print_image_info(hDC, x_pos, y_pos, self.barcode_images[0])
                y_pos = self.print_image_info(hDC, x_pos, y_pos, self.qrcode_images[0])
            
            hDC.EndPage()
            hDC.EndDoc()
        except pywintypes.error as e:
            messagebox.showerror("打印错误", f"win32print相关错误:\n{str(e)}")
        except win32ui.error as e:
            messagebox.showerror("打印错误", f"win32ui相关错误:\n{str(e)}")
        except Exception as e:
            messagebox.showerror("打印错误", f"其他错误:\n{str(e)}")
        
    def print_all(self):
        """打印所有条码信息"""
        if not self.entry_widgets:
            messagebox.showwarning("警告", "没有可打印的条码")
            return

        # 获取要打印的总数量
        total = len(self.entry_widgets)
        if total == 0:
            messagebox.showwarning("警告", "没有可打印的数据")
            return

        # 确认对话框
        if messagebox.askyesno("确认", f"确定要打印全部 {total} 条数据吗?"):
            # 开始打印流程
            self.printing = True
            self._start_blink()
            self.status_label.config(text=f"开始打印... 0/{total}")

            # 在后台线程中执行批量打印
            threading.Thread(target=self._print_all_thread, daemon=True).start()
    
    def _start_blink(self):
        """开始打印指示灯闪烁"""
        if not self.printing:
            self.print_indicator.config(bg="red")
            return

        current_color = self.print_indicator.cget("bg")
        next_color = "yellow" if current_color == "red" else "red"
        self.print_indicator.config(bg=next_color)
        self.blink_id = self.root.after(500, self._start_blink)  # 每500ms切换一次颜色

    def _stop_blink(self):
        """停止打印指示灯闪烁"""
        if self.blink_id:
            self.root.after_cancel(self.blink_id)
            self.blink_id = None
        self.print_indicator.config(bg="red")

    def _print_current_with_barcodes(self):
        """打印当前项目内容（包含条码和二维码）"""
        try:
            if not self.entry_widgets or len(self.entry_widgets) == 0:
                self.root.after(0, lambda: messagebox.showwarning("警告", "没有可打印的数据"))
                return

            first_row = self.entry_widgets[0]
            if not first_row or len(first_row) < 6:
                self.root.after(0, lambda: messagebox.showwarning("警告", "数据不完整"))
                return

            # 获取打印机名称
            printer_name = self.printer_name.get()
            hprinter = win32print.OpenPrinter(printer_name)
            
            try:
                # 获取打印机DC
                hdc = win32ui.CreateDC()
                hdc.CreatePrinterDC(printer_name)
                hdc.StartDoc("Barcode Print")
                hdc.StartPage()
                
                # 设置打印参数
                hdc.SetMapMode(win32con.MM_TWIPS)  # 使用TWIPS单位（1/1440英寸）
                left_margin = 1000  # 左边距
                top_margin = 1000   # 上边距
                line_height = 300   # 行高
                current_y = top_margin
                
                # 打印公司名称
                company = self.company_name.get()
                hdc.TextOut(left_margin, current_y, f"项目1: {company}")
                current_y += line_height * 2
                
                # 打印主要信息
                po = first_row[0].get() if first_row[0].get() else "无PO"
                date = self.format_date(first_row[1].get()) if first_row[1].get() else "无日期"
                item_no = first_row[2].get() if first_row[2].get() else "无项目号"
                part_no = first_row[3].get() if first_row[3].get() else "无料号"
                quantity = first_row[4].get() if first_row[4].get() else "无数量"
                name = first_row[5].get() if first_row[5].get() else "无名称"
                
                hdc.TextOut(left_margin, current_y, f"PO: {po}")
                current_y += line_height
                hdc.TextOut(left_margin, current_y, f"日期: {date}")
                current_y += line_height
                hdc.TextOut(left_margin, current_y, f"项目号: {item_no}")
                current_y += line_height
                hdc.TextOut(left_margin, current_y, f"料号: {part_no}")
                current_y += line_height
                hdc.TextOut(left_margin, current_y, f"数量: {quantity}")
                current_y += line_height * 2
                hdc.TextOut(left_margin, current_y, f"名称: {name}")
                current_y += line_height * 2
                
                # 打印详细字段
                hdc.TextOut(left_margin, current_y, "日期")
                current_y += line_height
                hdc.TextOut(left_margin, current_y, date)
                current_y += line_height * 2
                
                hdc.TextOut(left_margin, current_y, "项目号")
                current_y += line_height
                hdc.TextOut(left_margin, current_y, item_no)
                current_y += line_height * 2
                
                hdc.TextOut(left_margin, current_y, "料号")
                current_y += line_height
                hdc.TextOut(left_margin, current_y, part_no)
                current_y += line_height * 2
                
                hdc.TextOut(left_margin, current_y, "数量")
                current_y += line_height
                hdc.TextOut(left_margin, current_y, quantity)
                current_y += line_height * 2
                
                hdc.TextOut(left_margin, current_y, "名称")
                current_y += line_height
                hdc.TextOut(left_margin, current_y, name)
                current_y += line_height * 2
                
                # 打印条码和二维码
                if self.barcode_images and len(self.barcode_images) > 0:
                    for img_label, img_data, _ in self.barcode_images[0]:
                        if img_data:
                            # 将图像保存为临时文件
                            temp_img_path = os.path.join(tempfile.gettempdir(), "temp_barcode.bmp")
                            img_data.save(temp_img_path)
                            
                            # 加载位图
                            bmp = win32ui.CreateBitmap()
                            bmp.LoadImage(0, temp_img_path, win32con.IMAGE_BITMAP)
                            
                            # 创建内存DC并选择位图
                            memdc = hdc.CreateCompatibleDC()
                            memdc.SelectObject(bmp)
                            
                            # 获取位图尺寸
                            width = bmp.GetInfo()[0]
                            height = bmp.GetInfo()[1]
                            
                            # 打印图像
                            hdc.StretchBlt(
                                left_margin, current_y,
                                width * 1440 // 72, height * 1440 // 72,  # 转换为TWIPS单位
                                memdc, 0, 0, width, height, win32con.SRCCOPY
                            )
                            current_y += height * 1440 // 72 + line_height
                            
                            # 清理资源
                            memdc.DeleteDC()
                            bmp.DeleteObject()
                            os.remove(temp_img_path)
                
                hdc.EndPage()
                hdc.EndDoc()
                
                # 更新状态
                self.root.after(0, lambda: self.status_label.config(text="打印完成"))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("错误", f"打印失败:\n{str(e)}"))
                self.root.after(0, lambda: self.status_label.config(text="打印失败"))
            finally:
                win32print.ClosePrinter(hprinter)
                if 'hdc' in locals():
                    hdc.DeleteDC()
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"打印失败:\n{str(e)}"))
            self.root.after(0, lambda: self.status_label.config(text="打印失败"))
        finally:
            self.root.after(0, lambda: setattr(self, 'printing', False))
            self.root.after(0, self._stop_blink)

    def _print_all_thread(self):
        """后台线程执行批量打印任务"""
        try:
            total = len(self.entry_widgets)
            success_count = 0
            failed_count = 0
            for i, row_entries in enumerate(self.entry_widgets):
                if len(row_entries) < 6:
                    failed_count += 1
                    continue
                # 提取数据
                po = row_entries[0].get() if row_entries[0].get() else "无PO"
                date = self.format_date(row_entries[1].get()) if row_entries[1].get() else "无日期"
                item_no = row_entries[2].get() if row_entries[2].get() else "无项目号"
                part_no = row_entries[3].get() if row_entries[3].get() else "无料号"
                quantity = row_entries[4].get() if row_entries[4].get() else "无数量"
                name = row_entries[5].get() if row_entries[5].get() else "无名称"
                print(f"[DEBUG] 打印第 {i + 1}/{total} 条数据: PO={po}, 日期={date}, 项目号={item_no}, 料号={part_no},数量={quantity},名称={name}")
                # 更新UI显示当前进度
                self.root.after(0, lambda idx=i + 1, tot=total: self.status_label.config(
                    text=f"正在打印... {idx}/{tot}"))
                # 构造打印内容
                print_content = (
                    f"=== 条码打印 ({i + 1}/{total}) ===\n\n"
                    f"PO: {po}\n"
                    f"日期: {date}\n"
                    f"项目号: {item_no}\n"
                    f"料号: {part_no}\n"
                    f"数量: {quantity}\n"
                    f"名称: {name}\n\n"
                    f"{self.company_name.get()}\n\n"
                )
                # 选择打印机并打印
                printer_name = self.printer_name.get()
                hPrinter = win32print.OpenPrinter(printer_name)
                try:
                    hJob = win32print.StartDocPrinter(hPrinter, 1, ("Barcode Print", None, "RAW"))
                    win32print.StartPagePrinter(hPrinter)
                    win32print.WritePrinter(hPrinter, print_content.encode('utf-8'))
                    win32print.EndPagePrinter(hPrinter)
                    win32print.EndDocPrinter(hPrinter)
                    success_count += 1
                except win32print.pywintypes.error as e:  # 修改：使用pywintypes.error
                    print(f"[ERROR] 打印第 {i + 1} 条数据失败: {str(e)}")
                    failed_count += 1
                except Exception as e:
                    print(f"[ERROR] 打印第 {i + 1} 条数据失败: {str(e)}")
                    failed_count += 1
                finally:
                    win32print.ClosePrinter(hPrinter)
                time.sleep(0.5)  # 每条打印间隔
            # 打印完成
            result_msg = f"打印完成！成功: {success_count}, 失败: {failed_count}"
            self.root.after(0, lambda: messagebox.showinfo("打印结果", result_msg))
            self.root.after(0, lambda: self.status_label.config(text=result_msg))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"批量打印失败:\n{str(e)}"))
            self.root.after(0, lambda: self.status_label.config(text="批量打印失败"))
        finally:
            self.root.after(0, lambda: setattr(self, 'printing', False))
            self.root.after(0, self._stop_blink)

    def on_closing(self):
        """窗口关闭时断开连接"""
        if self.blink_id:
            self.root.after_cancel(self.blink_id)
        self.root.destroy()

    def load_excel(self):
        """加载Excel文件"""
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx *.xls")])
        if self.file_path:
            try:
                self.workbook = load_workbook(self.file_path, data_only=True)
                sheet = self.workbook.active
                print("[DEBUG] 读取到的第一行数据:", sheet["A1"].value, sheet["B1"].value)
                self.status_label.config(text=f"已加载文件: {os.path.basename(self.file_path)}")
            except Exception as e:
                messagebox.showerror("错误", f"文件加载失败:\n{str(e)}")
                self.status_label.config(text="文件加载失败")
        else:
            self.status_label.config(text="未选择文件")

    def clear_display(self):
        """清除显示内容"""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.entry_widgets = []
        self.barcode_images = []
        self.qrcode_images = []
        self.all_images = []  # 新增：清空图像引用列表
        self.status_label.config(text="显示已清除")

    def format_date(self, value):
        """格式化日期为YYYY-MM-DD格式"""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        elif isinstance(value, str) and " " in value:
            # 处理包含时间的字符串
            return value.split()[0]
        return str(value) if value is not None else ""

    def generate_barcode_image(self, data, width=300, height=80):
        """生成一维条码图像"""
        try:
            if data is None or not str(data).strip():
                return None, None  # 返回PIL Image和PhotoImage

            data = self.format_date(data)
            data_str = str(data).encode('utf-8').decode('utf-8')
            
            code = barcode.get('code128', data_str, writer=ImageWriter())
            buffer = io.BytesIO()
            code.write(buffer, {
                'write_text': False,
                'module_height': 10,
                'quiet_zone': 2,
                'font_size': 0
            })
            
            buffer.seek(0)
            pil_img = Image.open(buffer)
            pil_img = pil_img.resize((width, height), Image.LANCZOS)
            
            # 返回PIL Image和PhotoImage
            return pil_img, ImageTk.PhotoImage(pil_img)
        except Exception as e:
            print(f"生成条码失败: {e}")
            return None, None

    def generate_qr_code_image(self, data, size=180):
        """生成二维码图像"""
        try:
            if data is None or not str(data).strip():
                return None, None  # 返回PIL Image和PhotoImage

            data = self.format_date(data)
            data_str = str(data).encode('utf-8').decode('utf-8')
            
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(data_str)
            qr.make(fit=True)
            
            pil_img = qr.make_image(fill_color="black", back_color="white")
            pil_img = pil_img.resize((size, size), Image.LANCZOS)
            
            # 返回PIL Image和PhotoImage
            return pil_img, ImageTk.PhotoImage(pil_img)
        except Exception as e:
            print(f"生成二维码失败: {e}")
            return None, None

    def generate_barcodes(self):
        """生成条码并显示在界面上"""
        if not self.workbook:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        try:
            sheet = self.workbook.active
            self.clear_display()
            # 列映射关系
            col_map = {
                "序号": "B",
                "PO": "C",
                "日期": "D",
                "项目号": "E",
                "数量": "F",
                "料号": "G",
                "名称": "H"
            }
            
            # 创建一个列表来保持所有图像的引用
            self.all_images = []  
            for row in range(3, sheet.max_row + 1):
                item_no = sheet[f'{col_map["序号"]}{row}'].value
                po = sheet[f'{col_map["PO"]}{row}'].value
                if not po:
                    continue
                # 创建卡片式容器
                card_frame = ttk.Frame(self.scrollable_frame, style="Card.TFrame")
                card_frame.pack(fill="x", pady=5, padx=5)
                # 顶部标题栏
                title_frame = tk.Frame(card_frame, bg="#eaeaea")
                title_frame.pack(fill="x", pady=(0, 5))
                # 显示项目标题
                item_name = sheet[f'{col_map["名称"]}{row}'].value or "未命名项目"
                tk.Label(title_frame,
                        text=f"项目 {item_no}:",
                        font=('Arial', 10, 'bold'),
                        bg="#eaeaea").pack(side="left", padx=5)
                # 公司名称输入框
                company_entry = ttk.Entry(title_frame, textvariable=self.company_name, font=('Arial', 10), width=30)
                company_entry.pack(side="left", padx=5)
                # 创建三列布局（信息+条码+料号和二维码）
                data_frame = tk.Frame(card_frame, bg="#f0f0f0")
                data_frame.pack(fill="x")
                # 第一列：文本信息
                info_frame = tk.Frame(data_frame, bg="#f0f0f0")
                info_frame.pack(side="left", padx=10, pady=5, fill="y")
                # 显示详细信息
                fields = [
                    ("PO", f'{col_map["PO"]}{row}'),
                    ("日期", f'{col_map["日期"]}{row}'),
                    ("项目号", f'{col_map["项目号"]}{row}'),
                    ("料号", f'{col_map["料号"]}{row}'),
                    ("数量", f'{col_map["数量"]}{row}'),
                    ("名称", f'{col_map["名称"]}{row}')
                ]
                for label, cell_ref in fields:
                    value = sheet[cell_ref].value
                    value = self.format_date(value)
                    tk.Label(info_frame,
                            text=f"{label}: {value}" if value else f"{label}: -",
                            anchor="w",
                            font=('Arial', 9),
                            bg="#f0f0f0").pack(fill="x")
                # 第二列：一维条码（除数量外）
                barcode_col1 = tk.Frame(data_frame, bg="#f0f0f0")
                barcode_col1.pack(side="left", fill="y", expand=True)
                # 第三列：数量条码和二维码
                part_no_qrcode_col = tk.Frame(data_frame, bg="#f0f0f0")
                part_no_qrcode_col.pack(side="left", fill="y", expand=True)
                # 列间隔
                spacer1 = tk.Frame(data_frame, width=60, bg="#f0f0f0")
                spacer1.pack(side="left", fill="y")
                spacer2 = tk.Frame(data_frame, width=60, bg="#f0f0f0")
                spacer2.pack(side="left", fill="y")
                # 一维条码字段（不包含数量）
                barcode_fields = [
                    ("PO", sheet[f'{col_map["PO"]}{row}'].value),
                    ("日期", sheet[f'{col_map["日期"]}{row}'].value),
                    ("项目号", sheet[f'{col_map["项目号"]}{row}'].value),
                    ("料号", sheet[f'{col_map["料号"]}{row}'].value)
                ]
                row_entry_widgets = []  # 存储当前行的输入框
                row_barcode_images = []  # 存储当前行的条码图像
                row_qrcode_images = []  # 存储当前行的二维码图像
                # 生成一维条码及输入框
                for label, data in barcode_fields:
                    # 创建条码容器（包含标签、输入框和条码）
                    barcode_item_frame = tk.Frame(barcode_col1, bg="#f0f0f0")
                    barcode_item_frame.pack(fill="x", pady=2)
                    # 标签和输入框的容器
                    label_entry_frame = tk.Frame(barcode_item_frame, bg="#f0f0f0")
                    label_entry_frame.pack(fill="x")
                    # 左侧标签
                    tk.Label(label_entry_frame, text=label,
                            font=('Arial', 8), bg="#f0f0f0").pack(side="left")
                    # 右侧输入框
                    entry = ttk.Entry(label_entry_frame, font=('Arial', 8), width=30)
                    entry.pack(side="left", padx=8)
                    entry.insert(0, data if data else "")
                    row_entry_widgets.append(entry)
                    # 生成条码图像
                    pil_img, barcode_img = self.generate_barcode_image(data, width=280, height=70)
                    if barcode_img:
                        img_label = tk.Label(barcode_item_frame, image=barcode_img, bg="#f0f0f0")
                        img_label.image = barcode_img  # 保持引用
                        self.all_images.append(barcode_img)  # 关键修改：将图像添加到全局引用列表
                        img_label.pack()
                        row_barcode_images.append((img_label, pil_img, data))
                    else:
                        row_barcode_images.append((None, None, data))
                # 生成数量条码和二维码
                part_no_data = sheet[f'{col_map["数量"]}{row}'].value
                name_data = sheet[f'{col_map["名称"]}{row}'].value
                # 数量条码容器
                part_no_frame = tk.Frame(part_no_qrcode_col, bg="#f0f0f0", padx=35)
                part_no_frame.pack(fill="x", pady=5)
                # 标签和输入框的容器
                part_no_label_entry_frame = tk.Frame(part_no_frame, bg="#f0f0f0")
                part_no_label_entry_frame.pack(fill="x")
                # 左侧标签
                tk.Label(part_no_label_entry_frame, text="数量",
                        font=('Arial', 8), bg="#f0f0f0").pack(side="left")
                # 右侧输入框
                part_no_entry = ttk.Entry(part_no_label_entry_frame, font=('Arial', 8), width=30)
                part_no_entry.pack(side="left", padx=20)
                part_no_entry.insert(0, part_no_data if part_no_data else "")
                row_entry_widgets.append(part_no_entry)
                # 生成数量条码图像
                pil_img, part_no_barcode_img = self.generate_barcode_image(part_no_data, width=280, height=70)
                if part_no_barcode_img:
                    part_no_img_label = tk.Label(part_no_frame, image=part_no_barcode_img, bg="#f0f0f0")
                    part_no_img_label.image = part_no_barcode_img  # 保持引用
                    self.all_images.append(part_no_barcode_img)  # 关键修改：将图像添加到全局引用列表
                    part_no_img_label.pack(pady=5)
                    row_barcode_images.append((part_no_img_label, pil_img, part_no_data))
                else:
                    row_barcode_images.append((None, None, part_no_data))
                # 创建二维码容器
                qrcode_frame = tk.Frame(part_no_qrcode_col, bg="#f0f0f0", padx=35)
                qrcode_frame.pack(fill="x", pady=5)
                # 标签和输入框的容器
                qr_label_entry_frame = tk.Frame(qrcode_frame, bg="#f0f0f0")
                qr_label_entry_frame.pack(fill="x")
                # 左侧标签
                tk.Label(qr_label_entry_frame, text="名称",
                        font=('Arial', 8), bg="#f0f0f0").pack(side="left")
                # 右侧输入框
                name_entry = ttk.Entry(qr_label_entry_frame, font=('Arial', 8), width=30)
                name_entry.pack(side="left", padx=8)
                name_entry.insert(0, name_data if name_data else "")
                row_entry_widgets.append(name_entry)
                # 生成二维码图像
                pil_img, qr_img = self.generate_qr_code_image(name_data, size=180)
                if qr_img:
                    qr_label = tk.Label(qrcode_frame, image=qr_img, bg="#f0f0f0")
                    qr_label.image = qr_img  # 保持引用
                    self.all_images.append(qr_img)  # 关键修改：将图像添加到全局引用列表
                    qr_label.pack(pady=5)
                    row_qrcode_images.append((qr_label, pil_img, name_data))
                else:
                    row_qrcode_images.append((None, None, name_data))
                # 保存当前行的所有控件引用
                self.entry_widgets.append(row_entry_widgets)
                self.barcode_images.append(row_barcode_images)
                self.qrcode_images.append(row_qrcode_images)        
            self.status_label.config(text=f"已生成 {sheet.max_row - 2} 个项目的条码。")
            print(f"[DEBUG] 已生成 {len(self.entry_widgets)} 条数据，首条数据: {[e.get() for e in self.entry_widgets[0]]}")
        except Exception as e:
            messagebox.showerror("错误", f"生成条码时出错:\n{str(e)}")
            self.status_label.config(text="生成条码失败")

    def update_barcodes(self):
        """根据输入框内容更新条码"""
        if not self.entry_widgets:
            messagebox.showwarning("警告", "没有可更新的条码")
            return

        try:
            updated_count = 0
            for i, (row_entries, row_barcode_imgs, row_qrcode_imgs) in enumerate(
                    zip(self.entry_widgets, self.barcode_images, self.qrcode_images)
            ):
                # 更新一维条码
                for j, (img_label, _, _) in enumerate(row_barcode_imgs):
                    if j < len(row_entries):
                        new_data = row_entries[j].get()
                        new_img = self.generate_barcode_image(new_data, width=280, height=70)
                        if new_img and img_label:
                            img_label.config(image=new_img)
                            img_label.image = new_img  # 保持引用
                            updated_count += 1

                # 更新二维码
                if row_qrcode_imgs and len(row_entries) > len(row_barcode_imgs):
                    name_entry = row_entries[-1]
                    name_img_label, _, _ = row_qrcode_imgs[0]
                    new_name_data = name_entry.get()
                    new_qr_img = self.generate_qr_code_image(new_name_data, size=180)
                    if new_qr_img and name_img_label:
                        name_img_label.config(image=new_qr_img)
                        name_img_label.image = new_qr_img  # 保持引用
                        updated_count += 1

            self.status_label.config(text=f"已更新 {updated_count} 个条码")

        except Exception as e:
            messagebox.showerror("错误", f"更新条码时出错:\n{str(e)}")
            self.status_label.config(text="更新条码失败")


if __name__ == "__main__":
    root = tk.Tk()
    app = CompleteBarcodeGenerator(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()