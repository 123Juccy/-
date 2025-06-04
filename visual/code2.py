import io
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import barcode
from barcode.writer import ImageWriter
import qrcode
from openpyxl import load_workbook
from datetime import datetime
import os
import threading
import time
import win32print
import win32api
import tempfile
import pywintypes
import win32con
import win32ui
from PIL import ImageWin
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


# 手动定义可能缺失的常量
if not hasattr(win32con, 'DM_SIZE'):
    win32con.DM_SIZE = 40  # DEVMODE结构的大小，适用于Windows XP及以上系统
if not hasattr(win32con, 'DMPAPER_USER'):
    win32con.DMPAPER_USER = 0  # 用户自定义纸张


class CompleteBarcodeGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("条码生成与打印工具")
        self.root.geometry("1100x750")
        self.root.minsize(1000, 600)

        # 打印状态相关变量
        self.printing = False
        self.blink_id = None  # 用于存储闪烁定时器ID

        # 设置样式
        self.style = ttk.Style()
        self.style.configure("Card.TFrame", background="#f0f0f0", borderwidth=1, relief="solid", padding=10)
        self.style.configure("Title.TLabel", font=('Arial', 10, 'bold'), background="#f0f0f0")
        self.style.configure("Normal.TLabel", font=('Arial', 9), background="#f0f0f0")

        # 公司名称变量
        self.default_company_name = "珠海XXXXXX技术有限公司"

        # 打印机名称变量
        self.printer_name = tk.StringVar()
        self.available_printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
        if self.available_printers:
            self.printer_name.set(self.available_printers[0][2])

        # 创建界面元素
        self.create_widgets()
        # 注册中文字体
        try:
            # 尝试使用系统自带的中文字体
            font_path = "C:/Windows/Fonts/simhei.ttf"  # 黑体
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('SimHei', font_path))
                self.default_font = 'SimHei'
            else:
                # 如果找不到黑体，尝试其他中文字体
                font_path = "C:/Windows/Fonts/msyh.ttf"  # 微软雅黑
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('MSYH', font_path))
                    self.default_font = 'MSYH'
                else:
                    self.default_font = 'Helvetica'
        except:
            self.default_font = 'Helvetica'

    def create_widgets(self):
        # 顶部控制面板
        control_frame = tk.Frame(self.root, bg="#f0f0f0", padx=10, pady=10)
        control_frame.pack(fill="x")

        # 文件和条码操作按钮
        tk.Button(control_frame, text="选择 Excel 文件", command=self.load_excel,
                  bg="#4CAF50", fg="white", relief="flat").pack(side="left", padx=5)
        tk.Button(control_frame, text="生成条码", command=self.generate_barcodes,
                  bg="#2196F3", fg="white", relief="flat").pack(side="left", padx=5)
        tk.Button(control_frame, text="清除", command=self.clear_display,
                  bg="#f44336", fg="white", relief="flat").pack(side="left", padx=5)
        tk.Button(control_frame, text="更新条码", command=self.update_barcodes,
                  bg="#FF9800", fg="white", relief="flat").pack(side="left", padx=5)
        # 新增预览PDF按钮
        tk.Button(control_frame, text="预览PDF", command=self.preview_pdf,
                  bg="#9C27B0", fg="white", relief="flat").pack(side="left", padx=5)
        # 打印操作按钮
        printer_menu = ttk.OptionMenu(control_frame, self.printer_name, *[printer[2] for printer in self.available_printers])
        printer_menu.pack(side="left", padx=5)
        tk.Button(control_frame, text="打印当前", command=self.print_current,
                  bg="#00BCD4", fg="white", relief="flat").pack(side="left", padx=5)
        tk.Button(control_frame, text="打印所有", command=self.print_all,
                  bg="#E91E63", fg="white", relief="flat").pack(side="left", padx=5)

        # 状态指示灯
        status_indicators = tk.Frame(control_frame, bg="#f0f0f0")
        status_indicators.pack(side="right", padx=10)

        tk.Label(status_indicators, text="打印状态:", bg="#f0f0f0").pack(side="left")
        self.print_indicator = tk.Label(status_indicators, text="", width=5, bg="red", relief="solid")
        self.print_indicator.pack(side="left")

        # 状态栏
        self.status_label = tk.Label(self.root, text="准备就绪", bd=1, relief="sunken",
                                     anchor="w", bg="#f0f0f0", fg="#333")
        self.status_label.pack(fill="x", padx=10, pady=(0, 5))

        # 主显示区域
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # 创建 Canvas 和 Scrollbar
        self.canvas = tk.Canvas(self.main_frame, bg="#f5f5f5")
        self.scrollbar = ttk.Scrollbar(self.main_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # 初始化变量
        self.file_path = ""
        self.workbook = None
        self.entry_widgets = []  # 存储输入框引用
        self.barcode_images = []  # 存储条码图像引用
        self.qrcode_images = []  # 存储二维码图像引用
        self.all_images = []  # 存储所有图像引用
        self.company_names = []  # 存储每个项目的公司名称

    def print_text_info(self, hDC, x_pos, y_pos, fields):
        for label, value in fields:
            text = f"{label}: {value}"
            hDC.TextOut(x_pos, y_pos, text)
            y_pos += 20 * 1440 / 72
        return y_pos

    def print_image_info(self, hDC, x_pos, y_pos, img_labels):
        for img_label in img_labels:
            if img_label[0]:
                img = img_label[1]
                img = img.convert("RGB")
                photo = ImageTk.PhotoImage(img)
                img_width, img_height = img.size
                hDC.StretchBlt(x_pos, y_pos, img_width * 1440 / 72, img_height * 1440 / 72,
                               photo.width(), photo.height(), 0, 0, win32con.SRCCOPY)
                y_pos += img_height * 1440 / 72 + 20 * 1440 / 72
        return y_pos

    def print_current(self):
        """打印当前选中的条码信息（包含条码和二维码）"""
        if not self.entry_widgets:
            messagebox.showwarning("警告", "没有可打印的条码")
            return

        # 开始打印流程
        self.printing = True
        self._start_blink()
        self.status_label.config(text="正在打印...")

        # 在后台线程中执行打印
        threading.Thread(target=self._print_current_pdf, daemon=True).start()

    def print_all(self):
        """打印所有条码信息"""
        if not self.entry_widgets:
            messagebox.showwarning("警告", "没有可打印的条码")
            return

        # 获取要打印的总数量
        total = len(self.entry_widgets)
        if total == 0:
            messagebox.showwarning("警告", "没有可打印的数据")
            return

        # 确认对话框
        if messagebox.askyesno("确认", f"确定要打印全部 {total} 条数据吗?"):
            # 开始打印流程
            self.printing = True
            self._start_blink()
            self.status_label.config(text=f"开始打印... 0/{total}")

            # 在后台线程中执行批量打印
            threading.Thread(target=self._print_all_thread, daemon=True).start()

    def _start_blink(self):
        """开始打印指示灯闪烁"""
        if not self.printing:
            self.print_indicator.config(bg="red")
            return

        current_color = self.print_indicator.cget("bg")
        next_color = "yellow" if current_color == "red" else "red"
        self.print_indicator.config(bg=next_color)
        self.blink_id = self.root.after(500, self._start_blink)  # 每500ms切换一次颜色

    def _stop_blink(self):
        """停止打印指示灯闪烁"""
        if self.blink_id:
            self.root.after_cancel(self.blink_id)
            self.blink_id = None
        self.print_indicator.config(bg="red")

    def _print_current_pdf(self):
        """使用PDF作为中间格式打印当前项目(包含输入框内容)"""
        try:
            if not self.entry_widgets or len(self.entry_widgets) == 0:
                self.root.after(0, lambda: messagebox.showwarning("警告", "没有可打印的数据"))
                return

            # 获取当前选中的项目索引
            selected_index = self._get_selected_index()
            if selected_index is None:
                self.root.after(0, lambda: messagebox.showwarning("警告", "请先选择一个项目"))
                return

            row_entries = self.entry_widgets[selected_index]
            if not row_entries or len(row_entries) < 6:
                self.root.after(0, lambda: messagebox.showwarning("警告", "数据不完整"))
                return

            # 创建临时PDF文件
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_file:
                pdf_path = tmp_file.name

            # 创建PDF文档
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter

            # 设置字体 - 使用注册的中文字体
            c.setFont(self.default_font, 12)

            # 定义三列布局的起始位置
            col1_margin = 260  # 第一列左侧边距
            col2_margin = width * 0.35  # 第二列左侧边距
            col3_margin = width * 0.65  # 第三列左侧边距
            y_position = height - 50  # 初始Y坐标

            # 绘制公司名称 - 修复这里，确保获取正确的公司名称
            if selected_index < len(self.company_names):
                company_name = self.company_names[selected_index].get()
            else:
                company_name = self.default_company_name
            c.drawString(col1_margin, y_position, f"公司: {company_name}")
            y_position -= 30

            ''''
            # 第一列: 文本信息
            text_fields = [
                ("PO", row_entries[0].get() or "无PO"),
                ("日期", row_entries[1].get() or "无日期"),
                ("项目号", row_entries[2].get() or "无项目号"),
                ("料号", row_entries[3].get() or "无料号"),
                ("数量", row_entries[4].get() or "无数量"),
                ("名称", row_entries[5].get() or "无名称")
            ]
            
            for label, value in text_fields:
                c.drawString(col1_margin, y_position, f"{label}: {value}")
                y_position -= 20

            # 重置Y坐标
            y_position = height - 80
            '''
            # 第二列: PO、日期、项目号
            if selected_index < len(self.barcode_images) and len(self.barcode_images[selected_index]) > 0:
                for j, (img_label, img_data, original_data) in enumerate(self.barcode_images[selected_index][:3]):  # 只取前3个条码(PO、日期、项目号)
                    if img_data:
                        # 打印条码标签和输入框内容
                        if j < len(row_entries):
                            input_text = row_entries[j].get() if row_entries[j].get() else "无内容"
                            label_text = ["PO", "日期", "项目号"][j] if j < 3 else "未知"
                            c.drawString(col2_margin, y_position, f"{label_text}: {input_text}")
                            y_position -= 20
                        
                        # 将PIL Image转换为临时文件
                        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                            img_path = tmp_img.name
                            img_data.save(img_path)

                        img_width, img_height = img_data.size
                        scale = 0.5  # 缩放因子
                        c.drawImage(img_path, col2_margin, y_position - (img_height * scale),
                                    width=img_width * scale, height=img_height * scale)
                        y_position -= (img_height * scale) + 40

                        # 删除临时图像文件
                        try:
                            os.unlink(img_path)
                        except:
                            pass

            # 第三列: 料号条码、数量条码和名称二维码
            y_position = height - 80
            
            # 料号条码(从原来的第二列移过来)
            if selected_index < len(self.barcode_images) and len(self.barcode_images[selected_index]) > 3:  # 确保有料号条码
                material_img_label, material_img_data, _ = self.barcode_images[selected_index][3]
                if material_img_data:
                    # 打印料号标签和内容
                    material_text = row_entries[3].get() if row_entries[3].get() else "无料号"
                    c.drawString(col3_margin, y_position, f"料号: {material_text}")
                    y_position -= 20
                    
                    # 将PIL Image转换为临时文件
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                        img_path = tmp_img.name
                        material_img_data.save(img_path)

                    img_width, img_height = material_img_data.size
                    scale = 0.5  # 缩放因子
                    c.drawImage(img_path, col3_margin, y_position - (img_height * scale),
                                width=img_width * scale, height=img_height * scale)
                    y_position -= (img_height * scale) + 25

                    # 删除临时图像文件
                    try:
                        os.unlink(img_path)
                    except:
                        pass

            # 数量条码
            if selected_index < len(self.barcode_images) and len(self.barcode_images[selected_index]) > 4:  # 确保有数量条码
                quantity_img_label, quantity_img_data, _ = self.barcode_images[selected_index][4]
                if quantity_img_data:
                    # 打印数量标签和内容
                    quantity_text = row_entries[4].get() if row_entries[4].get() else "无数量"
                    c.drawString(col3_margin, y_position, f"数量: {quantity_text}")
                    y_position -= 20
                    
                    # 将PIL Image转换为临时文件
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                        img_path = tmp_img.name
                        quantity_img_data.save(img_path)

                    img_width, img_height = quantity_img_data.size
                    scale = 0.5  # 缩放因子
                    c.drawImage(img_path, col3_margin, y_position - (img_height * scale),
                                width=img_width * scale, height=img_height * scale)
                    y_position -= (img_height * scale) + 25

                    # 删除临时图像文件
                    try:
                        os.unlink(img_path)
                    except:
                        pass

            # 名称二维码
            if selected_index < len(self.qrcode_images) and len(self.qrcode_images[selected_index]) > 0:
                qr_label, qr_data, _ = self.qrcode_images[selected_index][0]
                if qr_data:
                    # 打印名称标签和内容
                    name_text = row_entries[5].get() if row_entries[5].get() else "无名称"
                    c.drawString(col3_margin, y_position, f"名称: {name_text}")
                    y_position -= 20
                    
                    # 将PIL Image转换为临时文件
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                        img_path = tmp_img.name
                        qr_data.save(img_path)

                    img_width, img_height = qr_data.size
                    scale = 0.5  # 缩放因子
                    c.drawImage(img_path, col3_margin, y_position - (img_height * scale),
                                width=img_width * scale, height=img_height * scale)
                    y_position -= (img_height * scale) + 40

                    # 删除临时图像文件
                    try:
                        os.unlink(img_path)
                    except:
                        pass

            c.save()

            # 使用系统默认打印机打印PDF
            printer_name = self.printer_name.get()
            win32api.ShellExecute(0, "print", pdf_path, f'"{printer_name}"', ".", 0)

            # 更新状态
            self.root.after(0, lambda: self.status_label.config(text="打印完成"))
            
            # 同时打开PDF预览
            os.startfile(pdf_path)
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"打印失败:\n{str(e)}"))
            self.root.after(0, lambda: self.status_label.config(text="打印失败"))
        finally:
            self.root.after(0, lambda: setattr(self, 'printing', False))
            self.root.after(0, self._stop_blink)

    def _print_all_thread(self):
        """后台线程执行批量打印任务"""
        try:
            total = len(self.entry_widgets)
            success_count = 0
            failed_count = 0
            # 创建临时PDF文件
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_file:
                pdf_path = tmp_file.name
            # 创建PDF文档
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter
            # 设置字体 - 使用注册的中文字体
            c.setFont(self.default_font, 12)
            # 布局参数
            col1_margin = 260  # 第一列左侧边距
            col2_margin = width * 0.35  # 第二列左侧边距
            col3_margin = width * 0.65  # 第三列左侧边距
            barcode_height = 50         # 条码统一高度
            scale = 0.5                 # 缩放因子

            for i, row_entries in enumerate(self.entry_widgets):
                if len(row_entries) < 6:
                    failed_count += 1
                    continue

                # 设置字体 - 使用注册的中文字体
                c.setFont(self.default_font, 12)

                y_position = height - 50  # 初始Y坐标

                # 绘制公司名称 - 修复这里，确保获取正确的公司名称
                if i < len(self.company_names):
                    company_name = self.company_names[i].get()
                else:
                    company_name = self.default_company_name
                c.drawString(col1_margin, y_position, f"公司: {company_name}")
                y_position -= 30

                # 第二列: PO、日期、项目号
                if i < len(self.barcode_images) and len(self.barcode_images[i]) > 0:
                    for j, (img_label, img_data, original_data) in enumerate(self.barcode_images[i][:3]):  # 只取前3个条码(PO、日期、项目号)
                        if img_data:
                            # 打印条码标签和输入框内容
                            if j < len(row_entries):
                                input_text = row_entries[j].get() if row_entries[j].get() else "无内容"
                                label_text = ["PO", "日期", "项目号"][j] if j < 3 else "未知"
                                c.drawString(col2_margin, y_position, f"{label_text}: {input_text}")
                                y_position -= 20
                            # 将PIL Image转换为临时文件
                            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                                img_path = tmp_img.name
                                img_data.save(img_path)
                            img_width, img_height = img_data.size
                            c.drawImage(img_path, col2_margin, y_position - (img_height * scale),
                                        width=img_width * scale, height=img_height * scale)
                            y_position -= (img_height * scale) + 40
                            # 删除临时图像文件
                            try:
                                os.unlink(img_path)
                            except:
                                pass

                # 第三列: 料号条码、数量条码和名称二维码
                y_position = height - 80

                # 料号条码(从原来的第二列移过来)
                if i < len(self.barcode_images) and len(self.barcode_images[i]) > 3:  # 确保有料号条码
                    material_img_label, material_img_data, _ = self.barcode_images[i][3]
                    if material_img_data:
                        # 打印料号标签和内容
                        material_text = row_entries[3].get() if row_entries[3].get() else "无料号"
                        c.drawString(col3_margin, y_position, f"料号: {material_text}")
                        y_position -= 20
                        # 将PIL Image转换为临时文件
                        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                            img_path = tmp_img.name
                            material_img_data.save(img_path)
                        img_width, img_height = material_img_data.size
                        c.drawImage(img_path, col3_margin, y_position - (img_height * scale),
                                    width=img_width * scale, height=img_height * scale)
                        y_position -= (img_height * scale) + 25
                        # 删除临时图像文件
                        try:
                            os.unlink(img_path)
                        except:
                            pass

                # 数量条码
                if i < len(self.barcode_images) and len(self.barcode_images[i]) > 4:  # 确保有数量条码
                    quantity_img_label, quantity_img_data, _ = self.barcode_images[i][4]
                    if quantity_img_data:
                        # 打印数量标签和内容
                        quantity_text = row_entries[4].get() if row_entries[4].get() else "无数量"
                        c.drawString(col3_margin, y_position, f"数量: {quantity_text}")
                        y_position -= 20
                        # 将PIL Image转换为临时文件
                        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                            img_path = tmp_img.name
                            quantity_img_data.save(img_path)
                        img_width, img_height = quantity_img_data.size
                        c.drawImage(img_path, col3_margin, y_position - (img_height * scale),
                                    width=img_width * scale, height=img_height * scale)
                        y_position -= (img_height * scale) + 25
                        # 删除临时图像文件
                        try:
                            os.unlink(img_path)
                        except:
                            pass

                # 名称二维码
                if i < len(self.qrcode_images) and len(self.qrcode_images[i]) > 0:
                    qr_label, qr_data, _ = self.qrcode_images[i][0]
                    if qr_data:
                        # 打印名称标签和内容
                        name_text = row_entries[5].get() if row_entries[5].get() else "无名称"
                        c.drawString(col3_margin, y_position, f"名称: {name_text}")
                        y_position -= 20
                        # 将PIL Image转换为临时文件
                        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                            img_path = tmp_img.name
                            qr_data.save(img_path)
                        img_width, img_height = qr_data.size
                        c.drawImage(img_path, col3_margin, y_position - (img_height * scale),
                                    width=img_width * scale, height=img_height * scale)
                        y_position -= (img_height * scale) + 40
                        # 删除临时图像文件
                        try:
                            os.unlink(img_path)
                        except:
                            pass

                # 每页一个项目
                c.showPage()
                # 更新状态
                self.root.after(0, lambda idx=i+1: self.status_label.config(text=f"正在生成PDF... {idx}/{total}"))
                success_count += 1

            c.save()
            # 打印PDF
            printer_name = self.printer_name.get()
            win32api.ShellExecute(0, "print", pdf_path, f'"{printer_name}"', ".", 0)
            # 更新状态
            result_msg = f"打印完成！成功: {success_count}, 失败: {failed_count}"
            self.root.after(0, lambda: messagebox.showinfo("打印结果", result_msg))
            self.root.after(0, lambda: self.status_label.config(text=result_msg))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"批量打印失败:\n{str(e)}"))
            self.root.after(0, lambda: self.status_label.config(text="批量打印失败"))
        finally:
            self.root.after(0, lambda: setattr(self, 'printing', False))
            self.root.after(0, self._stop_blink)

    def on_closing(self):
        """窗口关闭时断开连接"""
        if self.blink_id:
            self.root.after_cancel(self.blink_id)
        self.root.destroy()

    def load_excel(self):
        """加载Excel文件"""
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx *.xls")])
        if self.file_path:
            try:
                self.workbook = load_workbook(self.file_path, data_only=True)
                sheet = self.workbook.active
                print("[DEBUG] 读取到的第一行数据:", sheet["A1"].value, sheet["B1"].value)
                self.status_label.config(text=f"已加载文件: {os.path.basename(self.file_path)}")
            except Exception as e:
                messagebox.showerror("错误", f"文件加载失败:\n{str(e)}")
                self.status_label.config(text="文件加载失败")
        else:
            self.status_label.config(text="未选择文件")

    def clear_display(self):
        """清除显示内容"""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.entry_widgets = []
        self.barcode_images = []
        self.qrcode_images = []
        self.all_images = []  # 清空图像引用列表
        self.company_names = []  # 清空公司名称列表
        self.status_label.config(text="显示已清除")

    def format_date(self, value):
        """格式化日期为YYYY-MM-DD格式"""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        elif isinstance(value, str) and " " in value:
            # 处理包含时间的字符串
            return value.split()[0]
        return str(value) if value is not None else ""

    def generate_barcode_image(self, data, width=300, height=80):
        """生成一维条码图像"""
        try:
            if data is None or not str(data).strip():
                return None, None  # 返回PIL Image和PhotoImage

            data = self.format_date(data)
            data_str = str(data).encode('utf-8').decode('utf-8')

            code = barcode.get('code128', data_str, writer=ImageWriter())
            buffer = io.BytesIO()
            code.write(buffer, {
                'write_text': False,
                'module_height': 10,
                'quiet_zone': 2,
                'font_size': 0
            })

            buffer.seek(0)
            pil_img = Image.open(buffer)
            pil_img = pil_img.resize((width, height), Image.LANCZOS)

            # 返回PIL Image和PhotoImage
            return pil_img, ImageTk.PhotoImage(pil_img)
        except Exception as e:
            print(f"生成条码失败: {e}")
            return None, None

    def generate_qr_code_image(self, data, size=180):
        """生成二维码图像"""
        try:
            if data is None or not str(data).strip():
                return None, None  # 返回PIL Image和PhotoImage

            data = self.format_date(data)
            data_str = str(data).encode('utf-8').decode('utf-8')

            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(data_str)
            qr.make(fit=True)

            pil_img = qr.make_image(fill_color="black", back_color="white")
            pil_img = pil_img.resize((size, size), Image.LANCZOS)

            # 返回PIL Image和PhotoImage
            return pil_img, ImageTk.PhotoImage(pil_img)
        except Exception as e:
            print(f"生成二维码失败: {e}")
            return None, None

    def generate_barcodes(self):
        """生成条码并显示在界面上"""
        if not self.workbook:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        try:
            sheet = self.workbook.active
            self.clear_display()
            # 列映射关系
            col_map = {
                "序号": "B",
                "PO": "C",
                "日期": "D",
                "项目号": "E",
                "数量": "F",
                "料号": "G",
                "名称": "H"
            }

            # 创建一个列表来保持所有图像的引用
            self.all_images = []
            for row in range(3, sheet.max_row + 1):
                item_no = sheet[f'{col_map["序号"]}{row}'].value
                po = sheet[f'{col_map["PO"]}{row}'].value
                if not po:
                    continue
                
                # 创建卡片式容器
                card_frame = ttk.Frame(self.scrollable_frame, style="Card.TFrame")
                card_frame.pack(fill="x", pady=5, padx=5)
                
                # 顶部标题栏
                title_frame = tk.Frame(card_frame, bg="#eaeaea")
                title_frame.pack(fill="x", pady=(0, 5))
                
                # 显示项目标题
                item_name = sheet[f'{col_map["名称"]}{row}'].value or "未命名项目"
                tk.Label(title_frame,
                         text=f"项目 {item_no}:",
                         font=('Arial', 10, 'bold'),
                         bg="#eaeaea").pack(side="left", padx=5)
                
                # 公司名称输入框
                company_var = tk.StringVar(value=self.default_company_name)
                company_entry = ttk.Entry(title_frame, textvariable=company_var, font=('Arial', 10), width=30)
                company_entry.pack(side="left", padx=5)
                self.company_names.append(company_var)  # 保存公司名称变量
                
                # 创建三列布局（信息+条码+料号和二维码）
                data_frame = tk.Frame(card_frame, bg="#f0f0f0")
                data_frame.pack(fill="x")
                
                # 第一列：文本信息
                info_frame = tk.Frame(data_frame, bg="#f0f0f0", width=200)
                info_frame.pack(side="left", padx=10, pady=5, fill="y")
                
                # 第二列：PO、日期、项目号、料号条码
                barcode_col = tk.Frame(data_frame, bg="#f0f0f0")
                barcode_col.pack(side="left", fill="y", expand=True)
                
                # 第三列：数量条码和名称二维码
                qrcode_col = tk.Frame(data_frame, bg="#f0f0f0")
                qrcode_col.pack(side="left", fill="y", expand=True)
                
                # 显示详细信息
                fields = [
                    ("PO", f'{col_map["PO"]}{row}'),
                    ("日期", f'{col_map["日期"]}{row}'),
                    ("项目号", f'{col_map["项目号"]}{row}'),
                    ("料号", f'{col_map["料号"]}{row}'),
                    ("数量", f'{col_map["数量"]}{row}'),
                    ("名称", f'{col_map["名称"]}{row}')
                ]
                
                row_entry_widgets = []  # 存储当前行的输入框
                row_barcode_images = []  # 存储当前行的条码图像
                row_qrcode_images = []  # 存储当前行的二维码图像
                
                for label, cell_ref in fields:
                    # 为每个字段创建独立的输入框
                    field_frame = tk.Frame(info_frame, bg="#f0f0f0")
                    field_frame.pack(fill="x", pady=2)
                    
                    tk.Label(field_frame, text=f"{label}:", 
                             font=('Arial', 9), bg="#f0f0f0").pack(side="left")
                    
                    value = sheet[cell_ref].value
                    value = self.format_date(value)
                    
                    entry = ttk.Entry(field_frame, font=('Arial', 9), width=20)
                    entry.pack(side="left", padx=5)
                    entry.insert(0, value if value else "")
                    row_entry_widgets.append(entry)
                
                # 生成PO、日期、项目号、料号条码（放在中间列）
                barcode_fields = [
                    ("PO", sheet[f'{col_map["PO"]}{row}'].value),
                    ("日期", sheet[f'{col_map["日期"]}{row}'].value),
                    ("项目号", sheet[f'{col_map["项目号"]}{row}'].value),
                    ("料号", sheet[f'{col_map["料号"]}{row}'].value)
                ]
                
                for label, data in barcode_fields:
                    # 创建条码容器
                    barcode_item_frame = tk.Frame(barcode_col, bg="#f0f0f0")
                    barcode_item_frame.pack(fill="x", pady=2)
                    
                    # 标签
                    tk.Label(barcode_item_frame, text=label,
                             font=('Arial', 8), bg="#f0f0f0").pack()
                    
                    # 生成条码图像
                    pil_img, barcode_img = self.generate_barcode_image(data, width=250, height=60)
                    if barcode_img:
                        img_label = tk.Label(barcode_item_frame, image=barcode_img, bg="#f0f0f0")
                        img_label.image = barcode_img
                        self.all_images.append(barcode_img)
                        img_label.pack()
                        row_barcode_images.append((img_label, pil_img, data))
                    else:
                        row_barcode_images.append((None, None, data))
                
                # 生成数量条码（放在第三列）
                quantity_data = sheet[f'{col_map["数量"]}{row}'].value
                quantity_frame = tk.Frame(qrcode_col, bg="#f0f0f0")
                quantity_frame.pack(fill="x", pady=5)
                
                # 标签
                tk.Label(quantity_frame, text="数量",
                         font=('Arial', 8), bg="#f0f0f0").pack()
                
                # 生成数量条码图像
                pil_img, quantity_barcode_img = self.generate_barcode_image(quantity_data, width=250, height=60)
                if quantity_barcode_img:
                    quantity_img_label = tk.Label(quantity_frame, image=quantity_barcode_img, bg="#f0f0f0")
                    quantity_img_label.image = quantity_barcode_img
                    self.all_images.append(quantity_barcode_img)
                    quantity_img_label.pack()
                    row_barcode_images.append((quantity_img_label, pil_img, quantity_data))
                else:
                    row_barcode_images.append((None, None, quantity_data))
                
                # 生成名称二维码（放在第三列）
                name_data = sheet[f'{col_map["名称"]}{row}'].value
                qrcode_frame = tk.Frame(qrcode_col, bg="#f0f0f0")
                qrcode_frame.pack(fill="x", pady=5)
                
                # 标签
                tk.Label(qrcode_frame, text="名称",
                         font=('Arial', 8), bg="#f0f0f0").pack()
                
                # 生成二维码图像
                pil_img, qr_img = self.generate_qr_code_image(name_data, size=150)
                if qr_img:
                    qr_label = tk.Label(qrcode_frame, image=qr_img, bg="#f0f0f0")
                    qr_label.image = qr_img
                    self.all_images.append(qr_img)
                    qr_label.pack()
                    row_qrcode_images.append((qr_label, pil_img, name_data))
                else:
                    row_qrcode_images.append((None, None, name_data))
                
                # 保存当前行的所有控件引用
                self.entry_widgets.append(row_entry_widgets)
                self.barcode_images.append(row_barcode_images)
                self.qrcode_images.append(row_qrcode_images)

            self.status_label.config(text=f"已生成 {sheet.max_row - 2} 个项目的条码。")
            print(f"[DEBUG] 已生成 {len(self.entry_widgets)} 条数据，首条数据: {[e.get() for e in self.entry_widgets[0]]}")
        except Exception as e:
            messagebox.showerror("错误", f"生成条码时出错:\n{str(e)}")
            self.status_label.config(text="生成条码失败")

    def update_barcodes(self):
        """根据输入框内容更新条码"""
        if not self.entry_widgets:
            messagebox.showwarning("警告", "没有可更新的条码")
            return

        try:
            updated_count = 0
            for i, (row_entries, row_barcode_imgs, row_qrcode_imgs) in enumerate(
                    zip(self.entry_widgets, self.barcode_images, self.qrcode_images)
            ):
                # 更新一维条码
                for j, (img_label, _, _) in enumerate(row_barcode_imgs):
                    if j < len(row_entries):
                        new_data = row_entries[j].get()
                        pil_img, new_img = self.generate_barcode_image(new_data, width=250, height=60)
                        if new_img and img_label:
                            img_label.config(image=new_img)
                            img_label.image = new_img  # 保持引用
                            row_barcode_imgs[j] = (img_label, pil_img, new_data)  # 更新存储的图像数据
                            updated_count += 1

                # 更新二维码
                if row_qrcode_imgs and len(row_entries) > 5:  # 确保有名称字段
                    name_entry = row_entries[5]
                    name_img_label, _, _ = row_qrcode_imgs[0]
                    new_name_data = name_entry.get()
                    pil_img, new_qr_img = self.generate_qr_code_image(new_name_data, size=150)
                    if new_qr_img and name_img_label:
                        name_img_label.config(image=new_qr_img)
                        name_img_label.image = new_qr_img  # 保持引用
                        row_qrcode_imgs[0] = (name_img_label, pil_img, new_name_data)  # 更新存储的图像数据
                        updated_count += 1

            self.status_label.config(text=f"已更新 {updated_count} 个条码")

        except Exception as e:
            messagebox.showerror("错误", f"更新条码时出错:\n{str(e)}")
            self.status_label.config(text="更新条码失败")

    def preview_pdf(self):
        """预览当前项目的PDF文件(与打印内容完全一致)"""
        try:
            if not self.entry_widgets or len(self.entry_widgets) == 0:
                messagebox.showwarning("警告", "没有可预览的数据")
                return

            # 获取当前选中的项目索引
            selected_index = self._get_selected_index()
            if selected_index is None:
                messagebox.showwarning("警告", "请先选择一个项目")
                return

            row_entries = self.entry_widgets[selected_index]
            if not row_entries or len(row_entries) < 6:
                messagebox.showwarning("警告", "数据不完整")
                return

            # 创建临时PDF文件
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_file:
                pdf_path = tmp_file.name

            # 创建PDF文档
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter

            # 设置字体 - 使用注册的中文字体
            c.setFont(self.default_font, 12)

            # 定义三列布局的起始位置
            col1_margin = 260  # 第一列左侧边距
            col2_margin = width * 0.35  # 第二列左侧边距
            col3_margin = width * 0.65  # 第三列左侧边距
            y_position = height - 50  # 初始Y坐标

            # 绘制公司名称 - 修复这里，确保获取正确的公司名称
            if selected_index < len(self.company_names):
                company_name = self.company_names[selected_index].get()
            else:
                company_name = self.default_company_name
            c.drawString(col1_margin, y_position, f"公司: {company_name}")
            y_position -= 30

            # 第二列: PO、日期、项目号
            if selected_index < len(self.barcode_images) and len(self.barcode_images[selected_index]) > 0:
                for j, (img_label, img_data, original_data) in enumerate(self.barcode_images[selected_index][:3]):  # 只取前3个条码(PO、日期、项目号)
                    if img_data:
                        # 打印条码标签和输入框内容
                        if j < len(row_entries):
                            input_text = row_entries[j].get() if row_entries[j].get() else "无内容"
                            label_text = ["PO", "日期", "项目号"][j] if j < 3 else "未知"
                            c.drawString(col2_margin, y_position, f"{label_text}: {input_text}")
                            y_position -= 20
                        
                        # 将PIL Image转换为临时文件
                        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                            img_path = tmp_img.name
                            img_data.save(img_path)

                        img_width, img_height = img_data.size
                        scale = 0.5  # 缩放因子
                        c.drawImage(img_path, col2_margin, y_position - (img_height * scale),
                                    width=img_width * scale, height=img_height * scale)
                        y_position -= (img_height * scale) + 40

                        # 删除临时图像文件
                        try:
                            os.unlink(img_path)
                        except:
                            pass

            # 第三列: 料号条码、数量条码和名称二维码
            y_position = height - 80
            
            # 料号条码(从原来的第二列移过来)
            if selected_index < len(self.barcode_images) and len(self.barcode_images[selected_index]) > 3:  # 确保有料号条码
                material_img_label, material_img_data, _ = self.barcode_images[selected_index][3]
                if material_img_data:
                    # 打印料号标签和内容
                    material_text = row_entries[3].get() if row_entries[3].get() else "无料号"
                    c.drawString(col3_margin, y_position, f"料号: {material_text}")
                    y_position -= 20
                    
                    # 将PIL Image转换为临时文件
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                        img_path = tmp_img.name
                        material_img_data.save(img_path)

                    img_width, img_height = material_img_data.size
                    scale = 0.5  # 缩放因子
                    c.drawImage(img_path, col3_margin, y_position - (img_height * scale),
                                width=img_width * scale, height=img_height * scale)
                    y_position -= (img_height * scale) + 25

                    # 删除临时图像文件
                    try:
                        os.unlink(img_path)
                    except:
                        pass

            # 数量条码
            if selected_index < len(self.barcode_images) and len(self.barcode_images[selected_index]) > 4:  # 确保有数量条码
                quantity_img_label, quantity_img_data, _ = self.barcode_images[selected_index][4]
                if quantity_img_data:
                    # 打印数量标签和内容
                    quantity_text = row_entries[4].get() if row_entries[4].get() else "无数量"
                    c.drawString(col3_margin, y_position, f"数量: {quantity_text}")
                    y_position -= 20
                    
                    # 将PIL Image转换为临时文件
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                        img_path = tmp_img.name
                        quantity_img_data.save(img_path)

                    img_width, img_height = quantity_img_data.size
                    scale = 0.5  # 缩放因子
                    c.drawImage(img_path, col3_margin, y_position - (img_height * scale),
                                width=img_width * scale, height=img_height * scale)
                    y_position -= (img_height * scale) + 25

                    # 删除临时图像文件
                    try:
                        os.unlink(img_path)
                    except:
                        pass

            # 名称二维码
            if selected_index < len(self.qrcode_images) and len(self.qrcode_images[selected_index]) > 0:
                qr_label, qr_data, _ = self.qrcode_images[selected_index][0]
                if qr_data:
                    # 打印名称标签和内容
                    name_text = row_entries[5].get() if row_entries[5].get() else "无名称"
                    c.drawString(col3_margin, y_position, f"名称: {name_text}")
                    y_position -= 20
                    
                    # 将PIL Image转换为临时文件
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                        img_path = tmp_img.name
                        qr_data.save(img_path)

                    img_width, img_height = qr_data.size
                    scale = 0.5  # 缩放因子
                    c.drawImage(img_path, col3_margin, y_position - (img_height * scale),
                                width=img_width * scale, height=img_height * scale)
                    y_position -= (img_height * scale) + 40

                    # 删除临时图像文件
                    try:
                        os.unlink(img_path)
                    except:
                        pass

            c.save()

            # 使用系统默认的PDF阅读器打开文件进行预览
            os.startfile(pdf_path)

        except Exception as e:
            messagebox.showerror("错误", f"预览PDF失败:\n{str(e)}")
    def _get_selected_index(self):
        """获取当前选中项目的索引"""
        # 这里简化处理，返回第一个可见项目的索引
        # 实际应用中可能需要更复杂的逻辑来确定用户选择了哪个项目
        if self.entry_widgets:
            return 0
        return None


if __name__ == "__main__":
    root = tk.Tk()
    app = CompleteBarcodeGenerator(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()